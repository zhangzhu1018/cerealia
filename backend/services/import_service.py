"""
客户导入服务
- 解析 Excel (.xlsx) / CSV 文件
- 智能字段映射（支持多种列名变体）
- 国家自动识别
- 重复检测（公司名+国家去重）
- 自动背景评分
"""
import re
import io
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from ..models import db, Customer, Country, CustomerType, BackgroundCheck
from ..services.scoring_service import BackgroundScoringEngine

# ─── 字段映射规则 ─────────────────────────────────────────────────────────────
# 支持的列名变体 → 目标字段
FIELD_ALIASES = {
    'company_name_en': [
        'company_name_en', 'company_name', 'company', 'name', 'name_en',
        'company name', 'company name (en)', 'company name english',
        'firmenname', 'nom de la société', 'название компании',
        '企业名称', '企业名称(英文)', '企业英文名',
    ],
    'company_name_local': [
        'company_name_local', 'name_local', 'name_cn', 'name_local',
        '公司名称', '企业名称(中文)', '企业中文名',
        'nom local', 'местное название',
    ],
    'country': [
        'country', 'country_name', 'country_name_en', 'nation', 'region',
        'pays', 'land', 'страна',
        '国家', '国家名称',
    ],
    'city': [
        'city', 'town', 'location', '地址', '城市',
    ],
    'website': [
        'website', 'url', 'web', 'site', 'web_url',
        '网站', '网址',
    ],
    'email': [
        'email', 'e-mail', 'mail', 'contact_email', 'contact e-mail',
        '邮箱', '电子邮件',
    ],
    'phone': [
        'phone', 'tel', 'telephone', 'phone_number', 'contact_phone',
        '电话', '联系电话', '手机',
    ],
    'customer_type': [
        'customer_type', 'type', 'business_type', 'industry', 'category',
        '客户类型', '业务类型',
    ],
    'founded_year': [
        'founded_year', 'founded', 'established', 'established_year',
        '成立年份', '创立年份', '注册年份',
    ],
    'employee_count': [
        'employee_count', 'employees', 'employee', 'staff', 'staff_count',
        '员工数量', '员工人数', '员工',
    ],
    'annual_revenue': [
        'annual_revenue', 'revenue', 'turnover', 'sales', 'yearly_revenue',
        '年收入', '营业额', '年销售额',
    ],
    'has_import_history': [
        'has_import_history', 'import_history', 'imported_before', 'has_import',
        '有进口历史', '进口经验',
    ],
    'import_frequency': [
        'import_frequency', 'import_freq', 'frequency',
        '进口频率', '进口频次',
    ],
    'typical_import_volume': [
        'typical_import_volume', 'import_volume', 'import_amount', 'volume',
        '进口量', '进口数量', '典型进口量',
    ],
    'current_suppliers': [
        'current_suppliers', 'suppliers', 'supplier', 'current_supplier',
        '当前供应商', '供应商', '现有供应商',
    ],
    'has_cites_license': [
        'has_cites_license', 'cites', 'cites_license', 'cites_cert',
        'CITES认证', 'CITES许可证', 'CITES证书',
    ],
    'has_haccp_cert': [
        'has_haccp_cert', 'haccp', 'haccp_cert', 'haccp_certificate',
        'HACCP认证', 'HACCP证书',
    ],
    'market_segment': [
        'market_segment', 'segment', 'market', 'market_position',
        '市场细分', '市场定位', '细分市场',
    ],
    'notes': [
        'notes', 'remark', 'remarks', 'comment', 'comments', 'description',
        '备注', '说明', '描述',
    ],
    'tags': [
        'tags', 'label', 'labels', 'category_tags',
        '标签', '分类标签',
    ],
}

# 国家名 → 国家ID 映射（用于 country_id 解析）
COUNTRY_ALIASES = {
    # English
    'france': 'FR', 'germany': 'DE', 'united kingdom': 'GB', 'uk': 'GB',
    'england': 'GB', 'great britain': 'GB', 'usa': 'US', 'united states': 'US',
    'japan': 'JP', 'uae': 'AE', 'united arab emirates': 'AE',
    'australia': 'AU', 'canada': 'CA', 'china': 'CN', 'italy': 'IT',
    'spain': 'ES', 'netherlands': 'NL', 'belgium': 'BE', 'switzerland': 'CH',
    'russia': 'RU', 'russian federation': 'RU',
    # Chinese
    '法国': 'FR', '德国': 'DE', '英国': 'GB', '美国': 'US', '日本': 'JP',
    '阿联酋': 'AE', '澳大利亚': 'AU', '加拿大': 'CA', '中国': 'CN',
    '意大利': 'IT', '西班牙': 'ES', '荷兰': 'NL', '比利时': 'BE', '瑞士': 'CH',
    '俄罗斯': 'RU',
    # French
    'français': 'FR', 'francès': 'FR', 'allemagne': 'DE', 'italie': 'IT',
    'espagne': 'ES', 'pays-bas': 'NL',
}

# 客户类型 → ID 映射
CUSTOMER_TYPE_ALIASES = {
    '进口商': 'IMPORTER', 'importer': 'IMPORTER', 'importers': 'IMPORTER',
    '批发商': 'WHOLESALER', 'wholesaler': 'WHOLESALER', 'wholesalers': 'WHOLESALER',
    '品牌商': 'BRAND', 'brand': 'BRAND', 'brands': 'BRAND',
    '米其林餐厅': 'MICHELIN', 'michelin': 'MICHELIN', 'michelin star': 'MICHELIN',
    '高端酒店': 'HOTEL', 'hotel': 'HOTEL', 'hotels': 'HOTEL',
    '零售商': 'RETAILER', 'retailer': 'RETAILER', 'retailers': 'RETAILER',
    'other': 'OTHER', '其他': 'OTHER',
}


def _normalize_key(val: str) -> str:
    """将列名标准化为小写+去特殊字符"""
    if not val:
        return ''
    val = str(val).strip().lower()
    val = re.sub(r'[^a-z0-9]', '_', val)
    val = re.sub(r'_+', '_', val).strip('_')
    return val


def _build_alias_map() -> dict:
    """构建从标准化别名到目标字段的映射"""
    alias_to_field = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            alias_to_field[_normalize_key(alias)] = field
    return alias_to_field


def _detect_columns(headers: list[str]) -> dict[str, int]:
    """
    根据表头自动检测字段映射。
    返回 {target_field: column_index}
    """
    alias_map = _build_alias_map()
    mapping = {}
    for idx, header in enumerate(headers):
        norm = _normalize_key(header)
        if norm in alias_map:
            field = alias_map[norm]
            if field not in mapping:  # 取第一个匹配
                mapping[field] = idx
    return mapping


def _parse_country_id(country_str: str) -> int | None:
    """从国家字符串解析 country_id"""
    if not country_str:
        return None
    s = str(country_str).strip().lower()

    # 直接查库
    country_obj = Country.query.filter(
        db.or_(
            Country.name_en.ilike(s),
            Country.name_cn.ilike(s),
            Country.code.ilike(s),
        )
    ).first()
    if country_obj:
        return country_obj.id

    # 查别名
    code = COUNTRY_ALIASES.get(s)
    if code:
        country_obj = Country.query.filter_by(code=code).first()
        if country_obj:
            return country_obj.id

    return None


def _parse_customer_type_id(type_str: str) -> int | None:
    """从客户类型字符串解析 customer_type_id"""
    if not type_str:
        return None
    s = str(type_str).strip()
    code = CUSTOMER_TYPE_ALIASES.get(s.lower())
    if code:
        ct = CustomerType.query.filter_by(type_code=code).first()
        if ct:
            return ct.id
    # 模糊匹配
    ct = CustomerType.query.filter(
        db.or_(
            CustomerType.type_name_en.ilike(f'%{s}%'),
            CustomerType.type_name_cn.ilike(f'%{s}%'),
            CustomerType.type_code.ilike(f'%{s}%'),
        )
    ).first()
    return ct.id if ct else None


def _parse_bool(val) -> bool | None:
    """解析布尔值"""
    if val is None or val == '':
        return None
    s = str(val).strip().lower()
    if s in ('true', '1', 'yes', '是', '有', '是', '✓', '✔', '✅', 'y'):
        return True
    if s in ('false', '0', 'no', '否', '无', '没有', 'n'):
        return False
    return None


def _parse_int(val) -> int | None:
    try:
        return int(float(str(val).replace(',', '').replace(' ', '')))
    except (ValueError, TypeError):
        return None


def _parse_float(val) -> float | None:
    try:
        s = str(val).replace(',', '').replace(' ', '').replace('$', '').replace('€', '').replace('£', '')
        return float(s)
    except (ValueError, TypeError):
        return None


def _parse_date(val) -> date | None:
    if not val:
        return None
    try:
        if isinstance(val, date):
            return val
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%Y-%m', '%d-%m-%Y']:
            try:
                return datetime.strptime(str(val).strip(), fmt).date()
            except ValueError:
                continue
    except Exception:
        pass
    return None


def _row_to_dict(row: list, mapping: dict, headers: list) -> dict:
    """将一行数据转换为字典（只含已知字段）"""
    data = {}
    for field, idx in mapping.items():
        if idx < len(row):
            val = row[idx]
            # 特殊处理
            if field == 'country':
                data['country_id'] = _parse_country_id(val)
            elif field == 'customer_type':
                data['customer_type_id'] = _parse_customer_type_id(val)
            elif field in ('founded_year', 'employee_count'):
                data[field] = _parse_int(val)
            elif field == 'annual_revenue':
                data[field] = _parse_float(val)
            elif field == 'has_import_history':
                data[field] = _parse_bool(val)
            elif field == 'import_frequency':
                # 支持 "12次/年" "monthly" "quarterly" 等
                freq_map = {'monthly': 12, 'quarterly': 4, 'annually': 1, 'yearly': 1}
                s = str(val).lower().strip()
                if s in freq_map:
                    data[field] = freq_map[s]
                else:
                    data[field] = _parse_int(val)
            elif field == 'typical_import_volume':
                data[field] = _parse_float(val)
            elif field in ('has_cites_license', 'has_haccp_cert'):
                data[field] = _parse_bool(val)
            elif field == 'notes':
                data[field] = str(val) if val else None
            else:
                data[field] = str(val).strip() if val not in (None, '') else None
    return data


def _run_background_check(customer_data: dict, bg_data: dict) -> tuple[Customer, BackgroundCheck | None]:
    """
    创建客户记录并运行背景评分。
    返回 (customer_obj, bg_check_obj)
    """
    # 分离客户字段和背调字段
    customer_fields = {
        'company_name_en', 'company_name_local', 'country_id', 'city',
        'website', 'email', 'phone', 'customer_type_id',
        'follow_up_status', 'priority_level', 'notes', 'tags',
        'search_source', 'created_by',
    }
    customer_kwargs = {k: v for k, v in customer_data.items() if k in customer_fields and v is not None}
    # 默认值
    customer_kwargs.setdefault('follow_up_status', 'NEW')
    customer_kwargs.setdefault('priority_level', 'MEDIUM')

    # 背调数据
    bg_kwargs = {
        'founded_year': bg_data.get('founded_year'),
        'employee_count': bg_data.get('employee_count'),
        'annual_revenue': bg_data.get('annual_revenue'),
        'has_import_history': bg_data.get('has_import_history'),
        'import_frequency': bg_data.get('import_frequency'),
        'typical_import_volume': bg_data.get('typical_import_volume'),
        'current_suppliers': bg_data.get('current_suppliers'),
        'has_cites_license': bg_data.get('has_cites_license'),
        'has_haccp_cert': bg_data.get('has_haccp_cert'),
        'market_segment': bg_data.get('market_segment'),
        'scrape_date': date.today(),
    }
    # 清理 None
    bg_kwargs = {k: v for k, v in bg_kwargs.items() if v is not None}

    # 创建客户
    customer = Customer(**customer_kwargs)
    db.session.add(customer)
    db.session.flush()  # 获取 ID

    # 运行背景评分
    bg_check = None
    if bg_kwargs:
        bg_check = BackgroundCheck(customer_id=customer.id, **bg_kwargs)
        db.session.add(bg_check)
        db.session.flush()

        # 计算评分
        engine = BackgroundScoringEngine()
        score_result = engine.calculate_as_dict(bg_kwargs)

        # 更新客户评分字段
        customer.background_score = score_result['total_score']
        customer.import_trade_score = score_result['import_trade_score']
        customer.company_scale_score = score_result['company_scale_score']
        customer.market_position_score = score_result['market_position_score']
        customer.qualification_score = score_result['qualification_score']
        customer.cooperation_potential_score = score_result['cooperation_potential_score']
        customer.social_media_score = score_result['social_media_score']
        customer.responsiveness_score = score_result['responsiveness_score']

    return customer, bg_check


def parse_file(file_content: bytes, filename: str) -> tuple[list[list], list[str]]:
    """
    解析 Excel/CSV 文件。
    返回 (data_rows, headers)
    """
    import openpyxl
    import pandas as pd

    ext = filename.lower().split('.')[-1]

    if ext == 'csv':
        # 用 pandas 读取 CSV（自动处理编码）
        try:
            df = pd.read_csv(io.BytesIO(file_content), encoding='utf-8')
        except UnicodeDecodeError:
            df = pd.read_csv(io.BytesIO(file_content), encoding='latin-1')
        # 清理空行
        df = df.dropna(how='all')
        headers = list(df.columns)
        rows = df.fillna('').values.tolist()
    elif ext in ('xlsx', 'xls'):
        # 用 openpyxl 读取 Excel
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(cell) if cell is not None else '' for cell in row])
    else:
        raise ValueError(f'不支持的文件格式: {ext}，请上传 .xlsx 或 .csv 文件')

    return rows, headers


def import_customers(
    file_content: bytes,
    filename: str,
    run_background_check: bool = True,
    skip_duplicates: bool = True,
    created_by: str = 'import',
) -> dict:
    """
    批量导入客户主函数。

    返回摘要:
    {
        total: int,          # 总行数
        imported: int,       # 成功导入
        skipped: int,        # 跳过（缺必填/重复）
        failed: int,         # 失败
        errors: [str],       # 错误信息（最多20条）
        results: [{company_name_en, country, status, reason, customer_id}],
        country_stats: {country_name: count},
        grade_stats: {grade: count},
    }
    """
    # 1. 解析文件
    try:
        rows, headers = parse_file(file_content, filename)
    except Exception as e:
        raise ValueError(f'文件解析失败: {e}')

    if not rows:
        raise ValueError('文件中没有数据行')

    # 2. 检测列映射
    mapping = _detect_columns(headers)
    company_name_idx = mapping.get('company_name_en')
    country_idx = mapping.get('country')

    if company_name_idx is None:
        raise ValueError(
            f'未检测到「公司名称」列（支持: company_name_en, company, name 等）。'
            f'当前表头: {headers[:10]}'
        )

    # 3. 构建去重集合（仅按公司名去重，不区分国家；同公司名跨国家视为重复）
    existing = set()
    if skip_duplicates:
        existing_customers = db.session.query(
            Customer.company_name_en
        ).all()
        existing = {c.company_name_en.strip().lower() for c in existing_customers}

    # 4. 逐行处理
    imported = 0
    skipped = 0
    failed = 0
    errors = []
    results = []
    country_stats = {}
    grade_stats = {'A': 0, 'B': 0, 'C': 0, 'D': 0, 'E': 0, 'N': 0}

    for row_idx, row in enumerate(rows):
        row_num = row_idx + 2  # Excel行号（1是表头）

        try:
            # 提取数据
            data = _row_to_dict(row, mapping, headers)

            # 必填校验
            company_name = data.get('company_name_en', '').strip()
            if not company_name:
                skipped += 1
                results.append({
                    'row': row_num,
                    'company_name_en': '(空)',
                    'status': 'skipped',
                    'reason': '公司名称为空',
                })
                continue

            country_id = data.get('country_id')
            if not country_id:
                skipped += 1
                results.append({
                    'row': row_num,
                    'company_name_en': company_name,
                    'status': 'skipped',
                    'reason': f'无法识别国家（行内容: {str(row)[:80]}...）',
                })
                continue

            # 去重（仅按公司名，同名公司无论来自哪个国家均视为重复）
            if skip_duplicates and company_name.lower() in existing:
                skipped += 1
                results.append({
                    'row': row_num,
                    'company_name_en': company_name,
                    'country_id': country_id,
                    'status': 'skipped',
                    'reason': '客户已存在（公司名+国家重复）',
                })
                continue

            # 背景数据
            bg_data = {}
            if run_background_check:
                for fk in ('founded_year', 'employee_count', 'annual_revenue',
                           'has_import_history', 'import_frequency', 'typical_import_volume',
                           'current_suppliers', 'has_cites_license', 'has_haccp_cert',
                           'market_segment'):
                    if fk in data:
                        bg_data[fk] = data.pop(fk)
                bg_data['description'] = data.get('notes', '')

            # 补充客户数据
            data['created_by'] = created_by
            # 背景评分
            if run_background_check and bg_data:
                engine = BackgroundScoringEngine()
                score_result = engine.calculate_as_dict(bg_data)
                data['background_score'] = score_result['total_score']
                data['import_trade_score'] = score_result['import_trade_score']
                data['company_scale_score'] = score_result['company_scale_score']
                data['market_position_score'] = score_result['market_position_score']
                data['qualification_score'] = score_result['qualification_score']
                data['cooperation_potential_score'] = score_result['cooperation_potential_score']
                data['social_media_score'] = score_result['social_media_score']
                data['responsiveness_score'] = score_result['responsiveness_score']
                grade = score_result['grade']
            else:
                grade = 'N'

            # 写入数据库
            customer_kwargs = {
                'company_name_en': company_name,
                'company_name_local': data.get('company_name_local'),
                'country_id': country_id,
                'city': data.get('city'),
                'website': data.get('website'),
                'email': data.get('email'),
                'phone': data.get('phone'),
                'customer_type_id': data.get('customer_type_id'),
                'background_score': data.get('background_score', 0),
                'import_trade_score': data.get('import_trade_score', 0),
                'company_scale_score': data.get('company_scale_score', 0),
                'market_position_score': data.get('market_position_score', 0),
                'qualification_score': data.get('qualification_score', 0),
                'cooperation_potential_score': data.get('cooperation_potential_score', 0),
                'social_media_score': data.get('social_media_score', 0),
                'responsiveness_score': data.get('responsiveness_score', 0),
                'follow_up_status': 'NEW',
                'priority_level': 'MEDIUM',
                'notes': data.get('notes'),
                'tags': data.get('tags'),
                'search_source': 'file_import',
                'created_by': created_by,
            }
            # 过滤 None
            customer_kwargs = {k: v for k, v in customer_kwargs.items() if v is not None}

            customer = Customer(**customer_kwargs)
            db.session.add(customer)
            db.session.flush()

            # 背调记录
            if run_background_check and bg_data:
                bg_kwargs = {
                    'customer_id': customer.id,
                    'scrape_date': date.today(),
                }
                for fk in ('founded_year', 'employee_count', 'annual_revenue',
                           'has_import_history', 'import_frequency', 'typical_import_volume',
                           'current_suppliers', 'has_cites_license', 'has_haccp_cert', 'market_segment'):
                    if fk in bg_data and bg_data[fk] is not None:
                        bg_kwargs[fk] = bg_data[fk]
                if bg_kwargs:
                    bc = BackgroundCheck(**bg_kwargs)
                    db.session.add(bc)

            db.session.commit()

            # 统计
            imported += 1
            existing.add(company_name.lower())
            country_name = Country.query.get(country_id).name_en if country_id else 'Unknown'
            country_stats[country_name] = country_stats.get(country_name, 0) + 1
            grade_stats[grade] = grade_stats.get(grade, 0) + 1

            results.append({
                'row': row_num,
                'company_name_en': company_name,
                'country': country_name,
                'status': 'imported',
                'grade': grade,
                'score': float(data.get('background_score', 0)),
                'customer_id': customer.id,
            })

        except Exception as e:
            db.session.rollback()
            failed += 1
            err_msg = str(e)[:100]
            if len(errors) < 20:
                errors.append(f'第{row_num}行失败: {err_msg}')
            results.append({
                'row': row_num,
                'company_name_en': str(row[company_name_idx])[:50] if company_name_idx is not None else '(空)',
                'status': 'failed',
                'reason': err_msg,
            })

    return {
        'total': len(rows),
        'imported': imported,
        'skipped': skipped,
        'failed': failed,
        'errors': errors,
        'results': results,
        'country_stats': country_stats,
        'grade_stats': grade_stats,
    }


def preview_file(file_content: bytes, filename: str, max_rows: int = 5) -> dict:
    """
    预览文件前5行 + 自动检测的列映射。
    用于前端展示列映射 UI。
    """
    rows, headers = parse_file(file_content, filename)
    mapping = _detect_columns(headers)

    # 标准目标字段列表
    TARGET_FIELDS = [
        ('company_name_en', '公司英文名 *'),
        ('company_name_local', '公司本地名'),
        ('country', '国家 *'),
        ('city', '城市'),
        ('website', '网站'),
        ('email', '邮箱'),
        ('phone', '电话'),
        ('customer_type', '客户类型'),
        ('founded_year', '成立年份'),
        ('employee_count', '员工数量'),
        ('annual_revenue', '年收入'),
        ('has_import_history', '有进口历史'),
        ('import_frequency', '进口频率'),
        ('typical_import_volume', '典型进口量'),
        ('current_suppliers', '当前供应商'),
        ('has_cites_license', 'CITES认证'),
        ('has_haccp_cert', 'HACCP认证'),
        ('market_segment', '市场细分'),
        ('notes', '备注'),
    ]

    column_preview = []
    for field, label in TARGET_FIELDS:
        col_idx = mapping.get(field)
        sample_values = []
        if col_idx is not None:
            for row in rows[:max_rows]:
                if col_idx < len(row) and row[col_idx]:
                    sample_values.append(str(row[col_idx])[:40])
        column_preview.append({
            'field': field,
            'label': label,
            'detected': col_idx is not None,
            'column_index': col_idx,
            'source_header': headers[col_idx] if col_idx is not None else None,
            'sample_values': sample_values,
        })

    return {
        'headers': headers,
        'total_rows': len(rows),
        'column_preview': column_preview,
        'sample_rows': rows[:max_rows],
    }
